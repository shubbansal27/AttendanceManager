__author__ = 'Administrator'

import sys,os
import openpyxl
import sqlite3
import calendar
from PySide2.QtGui import *
from PySide2.QtWidgets import *
from GUI_Info import GUI_Info

class XLParser:

    def __init__(self,filename,dbname,frame):
        self.frame = frame
        self.filename = filename
        self.dbname = dbname

    def parse(self):

        self.returnVal = 1
        # creating workbook
        wb = openpyxl.load_workbook(self.filename)
        sheet = wb.get_active_sheet()


       # opening connection
        con = sqlite3.connect(self.dbname)


        #creating database & tables
        self.createDatabase(sheet,con)
        # check parsing success
        # if error occured, show error message and exit
        self.checkStatus(con)


        # extract data from excel & insert Attendance details
        self.insertEmployeeData(sheet,con)
        # check parsing success
        # if error occured, show error message and exit
        self.checkStatus(con)


        #commit changes
        con.commit()

        #closing connection
        con.close()


        #creating GUI_Info object
        objInfo = GUI_Info(self.dbname,self.month,self.year)
        # calculate info
        for empID in self.empIDList:
            objInfo.calculate(empID)



###########################################

    def insertEmployeeData(self,sheet,con):
        try:
            self.empIDList = list()
            ##getting holidays list
            query = 'select date from Holidays'
            holidayList = list()
            cursor = con.execute(query)
            for row in cursor:
                date = int(row[0].split('/')[1])
                holidayList.append(date)


            #getting employeeID, name
            for i in range(2,sheet.max_row,35):
                empID = sheet.cell(row=i,column=2).value
                empName = sheet.cell(row=i,column=6).value
                self.empIDList.append(empID)

                for k in range(1,31,1):
                    d_in = sheet.cell(row=i+1+k,column=5).value
                    d_out = sheet.cell(row=i+1+k,column=6).value

                    if d_in == None:
                        d_in = 'NA'
                    if d_out == None:
                        d_out = 'NA'

                    #removing white spaces
                    d_in = d_in.rstrip()
                    d_out = d_out.rstrip()

                    ##mark holiday
                    if 'NA' in d_in and k in holidayList:
                        d_in = 'Holiday'

                    if 'NA' in d_out and k in holidayList:
                        d_out = 'Holiday'


                    #inserting attendance in/out data
                    #in-attn
                    if(d_in != ''):
                        query = "update AttendanceIn set d" +str(k)+ "= '"+d_in + "' where empID='" + empID + "'";
                        con.execute(query)
                    #out-attn
                    if(d_out != ''):
                        query = "update AttendanceOut set d" +str(k)+ "= '"+d_out + "' where empID='" + empID + "'";
                        con.execute(query)

        except:
            self.showError()




    def createDatabase(self,sheet,con):

	print 'sheet = ', sheet
        try:
            ##extract basic details
            ##year,month, orgName
            date = sheet.cell(row=4,column=1).value
            self.year = int(date.split('-')[0])
            self.month = int(date.split('-')[1])
            orgName = sheet.cell(row=1,column=6).value
            ###create table
            query = 'create table BasicDetails(orgName varchar(100),year int, month int)';
            con.execute(query)
            #insert data
            query = "insert into BasicDetails values('" + orgName + "'," + str(self.year) + "," + str(self.month) + ")"
            con.execute(query)
           
            print "table created !!"
           ###creating table: AttendanceIn & AttendanceOut
            subquery = '(empID varchar(50) primary key, empName varchar(200)'
            for i in range(1,32,1):
                subquery = subquery + ',d' + str(i) + ' varchar(8)'

            subquery = subquery + ')'

            query = 'create table AttendanceIn'+subquery;
            con.execute(query)
            query = 'create table AttendanceOut'+subquery;
            con.execute(query)
           
            print "table created: AttendanceIn $$$$$$$$"
	     ##inserting employee details
            for i in range(2,sheet.max_row,35):
                empID = sheet.cell(row=i,column=2).value
                empName = sheet.cell(row=i,column=6).value
	        print empID, empName
                # insert EmployeeId & name in in/out tables
                query = "insert into AttendanceIn(empID,empName) values('" + empID + "','" + empName + "')"
                con.execute(query)
                query = "insert into AttendanceOut(empID,empName) values('" + empID + "','" + empName + "')"
                con.execute(query)

	    
	    print 'data entered:AttendanceOut !!'
	
            ###creating table: Holiday
            query = 'create table Holidays(date varchar(50) primary key,title varchar(150))'
            con.execute(query)
            #insert data
            #calculate 2nd and 4th saturday and sundays
            numDays = calendar.monthrange(self.year,self.month)[1]
            satCount = 0
	    print 'debug 1'
            for date in range(1,numDays+1,1):
                dayNo = calendar.weekday(self.year,self.month,date)
                if dayNo == 6:
                    query = "insert into Holidays values('"+ str(self.month) + '/' + str(date) + '/' + str(self.year) +"','Sunday')"
                    con.execute(query)
                elif dayNo == 5:
                    satCount = satCount + 1
                    if satCount == 2:
                        query = "insert into Holidays values('"+ str(self.month) + '/' + str(date) + '/' + str(self.year) +"','2nd Saturday')"
                        con.execute(query)
                    elif satCount == 4:
                        query = "insert into Holidays values('"+ str(self.month) + '/' + str(date) + '/' + str(self.year) +"','4th Saturday')"
                        con.execute(query)
	    	print 'debug 1: ', date, dayNo 



	    print 'data entered: Holidays'	
            # create Remarks/comments table
            ###creating table: AttendanceIn & AttendanceOut
            subquery = '(empID varchar(50) primary key'
            for i in range(1,32,1):
                subquery = subquery + ',d' + str(i) + ' text'

            subquery = subquery + ')'

            query = 'create table RemarkIn'+subquery;
            con.execute(query)
            query = 'create table RemarkOut'+subquery;
            con.execute(query)

	    print 'data entered: RemarksIn'
	
            ##insert employee-ID into both remark tables
            for i in range(2,sheet.max_row,35):
                empID = sheet.cell(row=i,column=2).value

                # insert EmployeeId
                query = "insert into RemarkIn(empID) values('" + empID + "')"
                con.execute(query)
                query = "insert into RemarkOut(empID) values('" + empID + "')"
                con.execute(query)


            # create table Calculations
            query = 'create table Calculations(empID varchar(50) primary key,NA float default 0,OD float default 0,Leave float default 0,WFH float default 0,Holiday float default 0,Extra float default 0,Late float default 0,Present float default 0,TotalDays float default 0)'
            con.execute(query)
            ##insert employee-ID
            for i in range(2,sheet.max_row,35):
                empID = sheet.cell(row=i,column=2).value
                query = "insert into Calculations(empID) values('" + empID + "')"
                con.execute(query)



	    print 'data entered: Calculations'	
            # commit all changes
            con.commit()

	    print 'data entered: commit statement '	
        except:
            self.showError()



#############################################################

    def showError(self):

        QMessageBox.critical(self.frame,'Error','Error occured while parsing excel.\n'+
                             'Please ensure you have imported valid biometric excel.\n\n' +
                             'Try again.')

        # set return Val to 0
        self.returnVal = 0


##############################################################

    def checkStatus(self,con):
       if self.returnVal == 0:

            # close connection, remove resource lock
            # in next step may need to delete localdb.db
            try:
                con.close()
            except:
                pass

            # remove localdb.db
            try:
                os.remove(self.dbname)
            except:
                pass

            # exit the application
            sys.exit()
