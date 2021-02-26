__author__ = 'admin'

from PySide2.QtGui import *
from PySide2.QtWidgets import *
import sqlite3
import calendar
import openpyxl
from openpyxl import cell
#from openpyxl.styles import Font, Style, Alignment,Color,PatternFill
from openpyxl.styles import Font, Alignment,Color,PatternFill

class Export:

    def __init__(self,dbname,month,year):
        self.dbname = dbname
        self.month = month
        self.year = year


    def exportToExcel(self,frame):

        # open save dialog
        r = QFileDialog.getSaveFileName(parent=frame,caption='Export',filter='*.xlsx')
        path = r[0]

        # if saved pressed then only
        if path != '':

            # opening connection
            con = sqlite3.connect(self.dbname)

            ##getting holidays dictionary
            query = 'select date,title from Holidays'
            holidays = dict()
            cursor = con.execute(query)
            for row in cursor:
                strTemp = row[0].split('/')
                date = strTemp[2]+'-'+strTemp[0]+'-'+strTemp[1]
                title = row[1]
                holidays[date] = title


            ## create Worksheet ##
            #wb_new = openpyxl.Workbook(guess_types=True)
            wb_new = openpyxl.Workbook()
            sheet = wb_new.get_active_sheet()
            sheet.title = 'data'
            sheet.freeze_panes = 'C3'
            alignDefault=Alignment(horizontal='center')
            fontDefault = Font(size=10)
            fontDates = Font(size=12,bold=True)
            #styleDefault = Style(font=fontDefault,alignment=alignDefault,number_format='general')
            #styleDates = Style(font=fontDates,alignment=alignDefault)

            #header
            sheet.cell(row=2,column=1).value = 'S.NO'
            sheet.cell(row=2,column=2).value = 'Name'
            sheet.column_dimensions['B'].width = 32
            for i in range(0,calendar.monthrange(self.year,self.month)[1],1):
                cValue = str(self.year)+'-'+str(self.month)+'-'+str(i+1)
                sheet.merge_cells(start_row=2,start_column=3+2*i,end_row=2,end_column=3+2*i+1)
                sheet.cell(row=2,column=3+2*i).value = cValue
                #sheet.cell(row=2,column=3+2*i).style = styleDates


            #Attendance-data
            queryIn = 'select * from AttendanceIn'
            cursorIn = con.execute(queryIn)
            queryOut = 'select * from AttendanceOut'
            cursorOut = con.execute(queryOut)
            j = 1
            for rowIn in cursorIn:
                rowOut = cursorOut.next()

                sheet.cell(row=j+2,column=1).value = str(j)
                #sheet.cell(row=j+2,column=1).style = styleDefault
                sheet.cell(row=j+2,column=2).value = rowIn[1]
                #sheet.cell(row=j+2,column=2).style = styleDefault

                #iterate over columns
                for k in range(0,calendar.monthrange(self.year,self.month)[1],1):
                    d_in = rowIn[k+2]
                    d_out = rowOut[k+2]

                    if d_in == None:
                        d_in = 'NA'

                    if d_out == None:
                        d_out = 'NA'

                    sheet.cell(row=j+2,column=3+2*k).value = d_in
                    sheet.cell(row=j+2,column=3+2*k+1).value = d_out

                    #style
                    #sheet.cell(row=j+2,column=3+2*k).style = styleDefault
                    #sheet.cell(row=j+2,column=3+2*k+1).style = styleDefault
                    sheet.cell(row=j+2,column=3+2*k).number_format = 'h:mm AM/PM'
                    sheet.cell(row=j+2,column=3+2*k+1).number_format = 'h:mm AM/PM'
                    #custom: cell color
                    self.colorCell(sheet.cell(row=j+2,column=3+2*k),d_in)
                    self.colorCell(sheet.cell(row=j+2,column=3+2*k+1),d_out)


                j = j + 1



            ## holidays ##
            max_row = sheet.max_row
            font1 = Font(size=12,bold=True)
            alignment1=Alignment(horizontal='center',
                                vertical='center',
                                text_rotation=90,
                                wrap_text=False,
                                shrink_to_fit=False,
                                indent=0)
            #styleObj1 = Style(font=font1,alignment=alignment1)

            for i in range(0,calendar.monthrange(self.year,self.month)[1],1):
                str_date = sheet.cell(row=2,column=3+2*i).value
                if str_date in holidays:
                    sheet.merge_cells(start_row=3,start_column=3+2*i,end_row=max_row,end_column=3+2*i+1)
                    ##sheet.cell(row=3,column=3+2*i).style = styleObj1
                    sheet.cell(row=3,column=3+2*i).value = holidays[str_date]



            ## Info/Calculations ##
            infoLabel = ['NA','OD','Leave','WFH','Holiday','Extra','Late','Present','TotalDays']
            query = 'select NA,OD,Leave,WFH,Holiday,Extra,Late,Present,TotalDays from Calculations'
            cursor = con.execute(query)
            j = 1
            max_col = sheet.max_column+1
            for row in cursor:
                for i in range(0,9,1):

                    #info header
                    if j == 1:
                        sheet.cell(row=j+1,column=max_col+i).value = infoLabel[i]
                        #sheet.cell(row=j+1,column=max_col+i).style = styleDates
                        self.colorCell(sheet.cell(row=j+1,column=max_col+i),infoLabel[i])

                    #info data
                    sheet.cell(row=j+2,column=max_col+i).value = row[i]
                    #sheet.cell(row=j+2,column=max_col+i).style = styleDefault

                j = j + 1


            ## save Worksheet ###
            wb_new.save(path)

            # closing connection
            con.close()


###################################################################

    def colorCell(self,cell,value):

        # ['NA','OD','Leave','WFH','Holiday','Extra','Late','Present','TotalDays']
        if 'NA' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('9bb7a7'))
        elif 'OD' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('ffe1ff'))
        elif 'Leave' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('fee773'))
        elif 'WFH' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('aaaaaa'))
        elif 'Holiday' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('9bb7a7'))
        elif 'Late' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('339966'))
        elif 'Extra' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('00ccff'))
        elif 'Present' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('ffcc5c'))
        elif 'TotalDays' in value:
            cell.fill=PatternFill(patternType='solid', fgColor=Color('ff6f69'))


